import os
import re
import shutil
import tempfile
import time
from pprint import pprint
from typing import List
from zipfile import ZipFile

import pandas as pd
import requests
from django.conf import settings
from django.core.management import BaseCommand
from openpyxl import load_workbook

config = {
    "default_header": 7,
    "special_file": [
        "TUSS - Demais terminologias - VERSÃO 202303.xlsx",
        "TUSS 64 - Envio de dados para ANS - Vers╞o 202303_Tab_18.xlsx",
        "TUSS 64 - Envio de dados para ANS - Vers╞o 202303_Tab_19_PARTE_1_DE_2.xlsx",
        "TUSS 64 - Envio de dados para ANS - Vers╞o 202303_Tab_19_PARTE_2_DE_2.xlsx"
        "TUSS 64 - Envio de dados para ANS - Vers╞o 202303_Tab_20.xlsx"
        "TUSS 64 - Envio de dados para ANS - Vers╞o 202303_Tab_22.xlsx",
    ],
    "special_file_header": 6,
    "ignore_sheets": ["Capa", "CAPA", "Indice"],
}


def move_csv_files(output_folder: str = "tuss_csvs"):
    root_dir = settings.MEDIA_ROOT
    output_dir = os.path.join(root_dir, output_folder)

    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Walk through the root directory and its subdirectories
    for subdir, _, files in os.walk(root_dir):
        for file in files:
            # Check if the file has a .csv extension
            if file.endswith(".csv"):
                file_path = os.path.join(subdir, file)
                # Move the file to the output directory
                shutil.move(file_path, os.path.join(output_dir, file))
                print(f"Moved {file_path} to {output_dir}")


def download_and_unzip():
    url = settings.TUSS_URL
    tuss_folder = settings.TUSS_FOLDER
    output_dir = os.path.join(settings.MEDIA_ROOT)

    if not tuss_folder in os.listdir(settings.MEDIA_ROOT):
        print("Downloading...")
        response = requests.get(url)
        if response.status_code != 200:
            raise Exception(f"Failed to download file from URL: {url}")

        # Save the zip file to a temporary directory
        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as temp_file:
            temp_file.write(response.content)
            temp_zip_path = temp_file.name

        # If no output directory is provided, create a directory inside the Django project root
        os.makedirs(output_dir, exist_ok=True)

        # Extract the zip file
        with ZipFile(temp_zip_path, "r") as zip_ref:
            zip_ref.extractall(output_dir)

        # Remove the downloaded zip file
        os.remove(temp_zip_path)

    # Walk through the extracted files and find nested zip files
    for subdir, _, files in os.walk(output_dir):
        for file in files:
            if file.endswith(".zip"):
                nested_zip_path = os.path.join(subdir, file)
                with ZipFile(nested_zip_path, "r") as nested_zip_ref:
                    nested_zip_ref.extractall(subdir)
                os.remove(nested_zip_path)

    return output_dir


def parse_xlsx_to_csv(root_path, config) -> List:
    converted_files = []

    for subdir, _, files in os.walk(root_path):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                file_path = os.path.join(subdir, file)
                workbook = load_workbook(file_path, read_only=True)
                table_name = file.split(" - ")[1].split(".")[0]

                for sheet_name in workbook.sheetnames:
                    if sheet_name not in config["ignore_sheets"]:
                        sheet = workbook[sheet_name]

                        # Search for the header row containing...
                        header_row_index = 1
                        for row in sheet.iter_rows(values_only=True):
                            if row[0] == "Código do Termo":
                                header = row
                                break
                            elif row[0] == "Código":
                                header = row
                                break
                            elif row[0] == "Código da Tabela":
                                header = row
                                break
                            elif row[0] == "Terminologia":
                                header = row
                                break

                            header_row_index += 1

                        # Read the data rows after the header
                        data = sheet.iter_rows(
                            min_row=header_row_index + 1, values_only=True
                        )

                        df = pd.DataFrame(data, columns=header)

                        # Remove any empty rows before the header
                        df.dropna(how="all", inplace=True)
                        df.reset_index(drop=True, inplace=True)

                        if file in config["special_file"]:
                            # Get the value from the 4th row, first column before creating the DataFrame
                            tab_number = int(sheet_name.split(" ")[1])

                            if tab_number > 74:
                                # table_title = sheet.cell(row=5, column=1).value
                                csv_file_name = f"{tab_number}.csv"

                            elif tab_number == 64:
                                if "PARTE" in sheet_name:
                                    table_title = sheet.cell(row=8, column=1).value
                                else:
                                    table_title = sheet.cell(row=4, column=1).value
                                csv_file_name = f"Tab {tab_number} - {table_title}).csv"

                            else:
                                # table_title = sheet.cell(row=4, column=1).value
                                csv_file_name = f"{tab_number}.csv"
                        else:
                            csv_file_name = f"{sheet_name}.csv"

                        table_number = int(
                            re.search(r"\d+", str(csv_file_name)).group()
                        )

                        if "MEDICAMENTOS" in csv_file_name:
                            table_number = 20

                        if "Materiais e OPME" in csv_file_name:
                            table_number = 19

                        if table_number != 87:
                            df["tabela"] = table_number

                        csv_file_path = os.path.join(subdir, csv_file_name)
                        df.to_csv(csv_file_path, index=False)
                        converted_files.append(csv_file_path)
                        print(f"Converted {csv_file_name}")

                workbook.close()

    move_csv_files()
    return converted_files


class Command(BaseCommand):
    help = "Download zip and convert to csv"

    def handle(self, *args, **kwargs):
        self.stdout.write(self.style.SUCCESS("Scraper starting..."))

        # Keep track of the time it took
        start_time = time.time()

        output_dir = download_and_unzip()

        converted_files = parse_xlsx_to_csv(output_dir, config)

        # Print the time it took
        print("--- %s seconds ---" % (time.time() - start_time))
        print(len(converted_files), "files converted")

        self.stdout.write(self.style.SUCCESS("Successfully converted files"))

        pprint(converted_files)
